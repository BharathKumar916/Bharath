import keyboard
import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select
class Base:
    #WebDriver:
    def browser_open(self):
        self.Chrome = webdriver. Chrome(executable_path=r"C:\Users\RANJITH\PycharmProjects\seleniumproject\Driver\chromedriver.exe")
        return self.Chrome
    def max_win (self):
        self.Chrome.maximize_window()
    def min_win(self):
        self.Chrome.minimize_window()
    def url_launch (self,url):
        self.Chrome.get(url)
    def wind_close(self):
        self.Chrome.close()
    def quit_browser(self):
        self.Chrome.quit()
    def screen_shot(self,location):
        self.Chrome.save_screenshot(location)
    def ex_script(self, data, element):
        self.Chrome.execute_script("arguments[0].setAttribute('value',{})".format(data),element)
    #WebElement
    def keys_insert(self,element,data):
        element.send_keys(data)
    def click_button(self,element):
        element.click()
    def att(self,element,att_name):
        pr_att = element.get_attribute(att_name)
        return pr_att
    #Select:
    def select_index(self,element,index):
        s =Select(element)
        s.select_by_index(index)
    def select_value(self,element,value):
        s = Select(element)
        s.select_by_value(value)
    def select_visibletext(self,element,text):
        s =Select(element)
        s .select_by_visible_text(text)
    def deselect_index(self,element,index):
        s =Select(element)
        s.select_by_index(index)
    def deselect_value(self,element,value):
        s1 = Select(element)
        s1.select_by_value(value)
    def deselect_visible_text(self,element,text):
        s2 =Select(element)
        s2.select_by_visible_text(text)
    def all_deselect(self,element,):
        s = Select(element)
        s.deselect_all()
    #ActionChains:
    def d_and_drop(self,src_element, des_element):
        a = ActionChains(self.Chrome)
        a.drag_and_drop(src_element,des_element).perform()
    def move_to(self, element):
        a = ActionChains(self.Chrome)
        a.move_to_element(element).perform()
    def r_click(self, element):
        a = ActionChains(self.Chrome)
        a.context_click(element).perform()
    def d_click(self,element):
        a = ActionChains(self.Chrome)
        a.double_click(element).perform()
    #alert
    def alert_accept(self):
        s =self.Chrome.switch_to.alert
        s.accept()
    def alert_dismiss(self):
        s =self.Chrome.switch_to.alert
        s.dismiss()
    def alert_insert(self,data):
        s =self.Chrome.switch_to.alert
        s.send_keys(data)
    def get_data_excel(self,row_index, column_index):
        loc =r"C:\Users\RANJITH\Desktop\HotelReg.xlsx"
        s = openpyxl.load_workbook(loc)
        active_sheet = s.active
        s1 = active_sheet.cell(row_index, column_index)
        data = s1.value
        return data
    # KeyBoard
    def press_down(self, times):
      for i in range(times):
        keyboard.press("down")
        keyboard.release("down")
    def press_enter(self,):
        keyboard.press("enter")
        keyboard.release("enter")
    def update_excel(self,val):
        loc = r"C:\Users\RANJITH\Desktop\amazon.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Data", 0)
        cell = sheet.cell(1, 1)
        cell.value = val
        workbook.save(loc)


